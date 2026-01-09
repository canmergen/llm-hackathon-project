# =============================================================================
# Banka Dokümanları Preparation Script - Excel Only
# =============================================================================
"""
Banka Dokümanları Preparation Script - Excel Only

This script processes Excel documents (XLSX) to extract,
chunk, and prepare vector-ready data for AI applications.

Features:
- Excel/XLSX document parsing
- Text chunking for retrieval systems
- Vector-ready subchunking with embedding optimization
- Comprehensive quality control and logging

Author: Can Mergen
Date: 2026-01-01
Version: 1.0.0
"""

# =============================================================================
# IMPORTS
# =============================================================================

import sys
import logging
from pathlib import Path
import re
import json
from dataclasses import dataclass, asdict
from scope_utils import assign_bank_document_scopes
from typing import List, Dict, Any, Optional
from collections import Counter

# Third-party imports
try:
    import openpyxl
except ImportError:
    openpyxl = None

# =============================================================================
# CONSTANTS
# =============================================================================

PROJECT_MARKERS = {"data", "src"}

# Chunking parameters
MAX_EMBEDDING_CHARS = 900
MIN_SUBCHUNK_CHARS = 80

# File paths (will be set after directory detection)
BANKA_DOKUMANLARI_RAW_DIR: Optional[Path] = None
BANKA_DOKUMANLARI_PARSED_DIR: Optional[Path] = None
BANKA_DOKUMANLARI_CHUNKS_DIR: Optional[Path] = None

# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class ExcelUnit:
    """Represents a parsed unit from an Excel document."""
    doc_id: str
    unit_id: str
    unit_type: str  # 'table_row'
    doc_type: str  # 'fee_schedule' or similar
    sheet_name: str
    row_number: int
    text: str
    metadata: Dict[str, Any]

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def find_project_root(start_path: Optional[Path] = None) -> Path:
    """Find the project root directory by searching for marker directories."""
    start_path = (start_path or Path.cwd()).resolve()
    for path in [start_path, *start_path.parents]:
        if PROJECT_MARKERS.issubset({item.name for item in path.iterdir() if item.is_dir()}):
            return path
    raise RuntimeError(f"Project root not found from: {start_path}")

def setup_logging(log_file_path: Path) -> logging.Logger:
    """Configure logging with both file and console handlers."""
    log_file_path.parent.mkdir(exist_ok=True)
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file_path, mode='w', encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    return logging.getLogger(__name__)

def setup_directories(base_dir: Path) -> None:
    """Set up global directory paths."""
    global BANKA_DOKUMANLARI_RAW_DIR, BANKA_DOKUMANLARI_PARSED_DIR, BANKA_DOKUMANLARI_CHUNKS_DIR
    
    banka_dokumanlari_dir = base_dir / "data" / "banka_dokumanlari"
    BANKA_DOKUMANLARI_RAW_DIR = banka_dokumanlari_dir / "raw"
    BANKA_DOKUMANLARI_PARSED_DIR = banka_dokumanlari_dir / "parsed"
    BANKA_DOKUMANLARI_CHUNKS_DIR = banka_dokumanlari_dir / "chunks"
    
    # Create directories if they don't exist
    BANKA_DOKUMANLARI_PARSED_DIR.mkdir(parents=True, exist_ok=True)
    BANKA_DOKUMANLARI_CHUNKS_DIR.mkdir(parents=True, exist_ok=True)

def safe_text(text: str) -> str:
    """Clean and normalize text for processing."""
    if not text:
        return ""
    text = text.replace("\r", "")
    return re.sub(r"\n{3,}", "\n\n", text).strip()

def normalize_for_embedding(text: str) -> str:
    """Normalize text for embedding by converting newlines to spaces."""
    if not text:
        return ""
    text = text.replace("\r", "")
    text = re.sub(r"\s*\n\s*", " ", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()

def determine_doc_type(doc_id: str) -> str:
    """Determine document type from doc_id."""
    doc_id_lower = doc_id.lower()
    if 'ucret' in doc_id_lower or 'tarife' in doc_id_lower:
        return 'fee_schedule'
    elif 'sozlesme' in doc_id_lower:
        return 'contract'
    return 'excel_data'

# =============================================================================
# EXCEL PARSER
# =============================================================================

def parse_excel_document(xlsx_path: Path, doc_id: str) -> List[Dict[str, Any]]:
    """Parse Excel document and extract units with meaningful sentences."""
    if openpyxl is None:
        raise ImportError("openpyxl is not installed. Install with: pip install openpyxl")

    logger.info(f"Parsing Excel: {xlsx_path}")
    units = []

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            logger.info(f"Processing sheet: {sheet_name}")

            # Extract headers from first row
            headers = []
            rows_data = list(sheet.iter_rows(values_only=True))

            if not rows_data:
                logger.warning(f"Empty sheet: {sheet_name}")
                continue

            # First row is header
            headers = [str(h).strip() if h is not None else f"Column{idx}" 
                      for idx, h in enumerate(rows_data[0], 1)]
            logger.info(f"Headers: {headers}")

            last_row_text = None
            # Process data rows (skip header)
            for row_idx, row in enumerate(rows_data[1:], 2):  # Start from row 2
                if not any(row):  # Skip empty rows
                    continue

                # Duplicate protection
                current_row_str = str(row)
                if current_row_str == last_row_text:
                    continue
                last_row_text = current_row_str

                # Create meaningful sentence from row data
                row_parts = []
                row_parts.extend(
                    f"{header}: {str(value).strip()}"
                    for header, value in zip(headers, row)
                    if value is not None and str(value).strip()
                )
                if not row_parts:
                    continue

                # Join parts to create a meaningful sentence
                row_text = ". ".join(row_parts) + "."

                unit = ExcelUnit(
                    doc_id=doc_id,
                    unit_id=f"{doc_id}_{sheet_name}_row_{row_idx}",
                    unit_type="table_row",
                    doc_type=determine_doc_type(doc_id),
                    sheet_name=sheet_name,
                    row_number=row_idx,
                    text=safe_text(row_text),
                    metadata={
                        "sheet": sheet_name, 
                        "row": row_idx, 
                        "num_columns": len(row),
                        "headers": headers
                    }
                )
                units.append(asdict(unit))

    except Exception as e:
        logger.error(f"Error parsing Excel {xlsx_path}: {e}")
        raise

    logger.info(f"Extracted {len(units)} units from Excel")
    return units

# =============================================================================
# SCOPE MAPPING
# =============================================================================

def guess_scope_from_text(text: str, doc_id: str) -> List[str]:
    """Map Excel row text to teblig benzeri scope etiketleri."""
    if not text:
        return ["whitelist_and_pricing"]
    
    text_lower = text.lower()
    # Ücret tarifeleri her zaman pricing'e girer
    scopes = {"whitelist_and_pricing"}
    
    # EFT / FAST / Havale
    if any(k in text_lower for k in ["eft", "fast", "havale", "para transfer", "swift"]):
        scopes.add("money_transfer")
        scopes.add("eft_havale_fast")
    
    # ATM
    if "atm" in text_lower:
        scopes.add("atm")
    
    # Kredi kartı
    if any(k in text_lower for k in ["kredi kart", "kart"]):
        scopes.add("credit_card")
    
    # Nakit avans
    if "nakit" in text_lower and "avans" in text_lower:
        scopes.add("credit_card_cash_advance")
    
    # Mevduat
    if any(k in text_lower for k in ["mevduat", "hesap", "vadesiz", "vadeli"]):
        scopes.add("deposit_account")
    
    # Dijital / elektronik / bilgilendirme
    if any(k in text_lower for k in ["elektronik", "dijital", "internet", "mobil", "sms", "bilgilendirme"]):
        scopes.add("disclosure_and_documentation")
    
    return sorted(scopes) if scopes else ["whitelist_and_pricing"]

# =============================================================================
# CHUNKING FUNCTIONS
# =============================================================================

def create_chunks(parsed_units: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Create chunks from parsed units."""
    chunks = []
    
    for unit in parsed_units:
        unit_id = unit.get("unit_id")
        if not unit_id:
            continue
        
        text = normalize_for_embedding(unit.get("text", ""))
        if not text:
            continue
        
        # Scope mapping
        scope = assign_bank_document_scopes(unit.get("doc_id", ""), text)
        
        chunk = {
            "chunk_id": unit_id,
            "doc_id": unit.get("doc_id"),
            "unit_id": unit_id,
            "unit_type": unit.get("unit_type"),
            "doc_type": unit.get("doc_type"),
            "sheet_name": unit.get("sheet_name"),
            "row_number": unit.get("row_number"),
            "text": text,
            "scope": scope,
            "metadata": unit.get("metadata", {})
        }
        chunks.append(chunk)
    
    logger.info(f"Created {len(chunks)} chunks")
    return chunks

def create_vector_records(chunk: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Create vector-ready records from a chunk. Returns ONLY child chunks."""
    full_text = chunk.get("text", "")
    if not full_text:
        return []
    
    parent_id = chunk["chunk_id"]
    records = []
    
    # If text is short enough, create single child
    if len(full_text) <= MAX_EMBEDDING_CHARS:
        child = dict(chunk)
        child["chunk_id"] = f"{parent_id}__c1"
        child["chunk_kind"] = "child"
        child["parent_chunk_id"] = parent_id
        child["text_for_embedding"] = full_text
        records.append(child)
        return records
    
    # Split long text by separator or sentences
    # First try splitting by | separator
    parts = [p.strip() for p in full_text.split('|') if p.strip()]
    
    if len(parts) > 1:
        # Multiple columns, create child for each column group
        for idx, part in enumerate(parts, 1):
            if len(part) >= MIN_SUBCHUNK_CHARS:
                child = dict(chunk)
                child["chunk_id"] = f"{parent_id}__col{idx}"
                child["chunk_kind"] = "child"
                child["parent_chunk_id"] = parent_id
                child["text_for_embedding"] = part[:MAX_EMBEDDING_CHARS]
                records.append(child)
    else:
        # Single long text, split by sentences
        sentences = re.split(r'(?<=[.!?])\s+', full_text)
        
        current_chunk = ""
        child_idx = 1
        
        for sentence in sentences:
            if len(current_chunk) + len(sentence) + 1 <= MAX_EMBEDDING_CHARS:
                current_chunk = f"{current_chunk} {sentence}".strip() if current_chunk else sentence
            else:
                if current_chunk:
                    child = dict(chunk)
                    child["chunk_id"] = f"{parent_id}__c{child_idx}"
                    child["chunk_kind"] = "child"
                    child["parent_chunk_id"] = parent_id
                    child["text_for_embedding"] = current_chunk
                    records.append(child)
                    child_idx += 1
                current_chunk = sentence
        
        # Add remaining chunk
        if current_chunk:
            child = dict(chunk)
            child["chunk_id"] = f"{parent_id}__c{child_idx}"
            child["chunk_kind"] = "child"
            child["parent_chunk_id"] = parent_id
            child["text_for_embedding"] = current_chunk
            records.append(child)
    
    return records

def perform_chunk_qc(chunks: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Perform quality control on chunks."""
    chunk_ids = [chunk.get("chunk_id") for chunk in chunks if chunk.get("chunk_id")]
    duplicate_ids = [k for k, v in Counter(chunk_ids).items() if v > 1]

    empty_texts = [chunk.get("chunk_id") for chunk in chunks if not (chunk.get("text") or "").strip()]

    text_lengths = [(chunk.get("chunk_id"), len(chunk.get("text", ""))) for chunk in chunks]
    sorted_lengths = sorted(text_lengths, key=lambda x: x[1], reverse=True)

    return {
        "counts": {
            "chunks": len(chunks),
            "sheets": len(
                {
                    chunk.get("sheet_name")
                    for chunk in chunks
                    if chunk.get("sheet_name")
                }
            ),
        },
        "duplicate_chunk_ids": duplicate_ids,
        "empty_text": empty_texts,
        "text_length_stats": {
            "min": min((length for _, length in sorted_lengths), default=0),
            "max": max((length for _, length in sorted_lengths), default=0),
            "top_10_longest": sorted_lengths[:10],
        },
    }

def perform_vector_qc(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Perform quality control on vector records."""
    chunk_ids = [record.get("chunk_id") for record in records if record.get("chunk_id")]
    duplicate_ids = [k for k, v in Counter(chunk_ids).items() if v > 1]
    
    children = [record for record in records if record.get("chunk_kind") == "child"]
    too_long = [record["chunk_id"] for record in children
               if len(record.get("text_for_embedding", "")) > MAX_EMBEDDING_CHARS]
    too_short = [record["chunk_id"] for record in children
                if 0 < len(record.get("text_for_embedding", "")) < MIN_SUBCHUNK_CHARS]
    missing_parent = [record["chunk_id"] for record in children
                     if not record.get("parent_chunk_id")]
    
    parents = {record["chunk_id"] for record in records if record.get("chunk_kind") == "parent"}
    orphan_children = [record["chunk_id"] for record in children
                      if record.get("parent_chunk_id") not in parents]
    
    return {
        "counts": {
            "records_total": len(records),
            "parents": len(parents),
            "children": len(children),
            "duplicate_chunk_ids": len(duplicate_ids),
            "child_too_long": len(too_long),
            "child_too_short": len(too_short),
            "child_missing_parent_id": len(missing_parent),
            "child_orphan": len(orphan_children),
        },
        "sample": {
            "duplicate_chunk_ids": duplicate_ids[:50],
            "child_too_long": too_long[:50],
            "child_too_short": too_short[:50],
            "child_missing_parent_id": missing_parent[:50],
            "child_orphan": orphan_children[:50],
        },
    }

# =============================================================================
# FILE I/O FUNCTIONS
# =============================================================================

def save_json(data: Any, file_path: Path) -> None:
    """Save data to JSON file."""
    try:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"Saved JSON to {file_path}")
    except Exception as e:
        logger.error(f"Error saving JSON to {file_path}: {e}")
        raise

def save_jsonl(data: List[Dict[str, Any]], file_path: Path) -> None:
    """Save list of dictionaries to JSONL file."""
    try:
        with open(file_path, "w", encoding="utf-8") as f:
            for item in data:
                f.write(json.dumps(item, ensure_ascii=False) + "\n")
        logger.info(f"Saved JSONL to {file_path}")
    except Exception as e:
        logger.error(f"Error saving JSONL to {file_path}: {e}")
        raise

# =============================================================================
# MAIN EXECUTION FUNCTIONS
# =============================================================================

def process_excel_document(file_path: Path) -> None:
    """Process a single Excel document."""
    logger.info(f"Processing Excel document: {file_path.name}")

    # Generate doc_id from filename
    doc_id = file_path.stem

    # Parse Excel
    units = parse_excel_document(file_path, doc_id)

    # Save parsed units
    parsed_output = BANKA_DOKUMANLARI_PARSED_DIR / f"{doc_id}_units.json" # pyright: ignore[reportOptionalOperand]
    save_json(units, parsed_output)

    # Create chunks
    chunks = create_chunks(units)
    _extracted_from_process_excel_document_17(
        doc_id, '_chunks.json', '_chunks.jsonl', chunks
    )
    # QC on chunks
    chunk_qc = perform_chunk_qc(chunks)
    chunk_qc_output = BANKA_DOKUMANLARI_CHUNKS_DIR / f"{doc_id}_chunks_qc.json" # pyright: ignore[reportOptionalOperand]
    save_json(chunk_qc, chunk_qc_output)

    # Create vector-ready records
    records = []
    for chunk in chunks:
        records.extend(create_vector_records(chunk))

    _extracted_from_process_excel_document_17(
        doc_id, '_vector_ready.json', '_vector_ready.jsonl', records
    )
    # QC on vector records
    vector_qc = perform_vector_qc(records)
    vector_qc_output = BANKA_DOKUMANLARI_CHUNKS_DIR / f"{doc_id}_vector_ready_qc.json" # pyright: ignore[reportOptionalOperand]
    save_json(vector_qc, vector_qc_output)

    logger.info(f"Completed processing: {file_path.name}")
    logger.info(f"  - Units: {len(units)}")
    logger.info(f"  - Chunks: {len(chunks)}")
    logger.info(f"  - Vector records: {len(records)}")
    logger.info(f"  - QC Summary: {chunk_qc['counts']}")


# TODO Rename this here and in `process_excel_document`
def _extracted_from_process_excel_document_17(doc_id, arg1, arg2, arg3):
    chunks_output = BANKA_DOKUMANLARI_CHUNKS_DIR / f"{doc_id}{arg1}" # pyright: ignore[reportOptionalOperand]
    chunks_jsonl_output = BANKA_DOKUMANLARI_CHUNKS_DIR / f"{doc_id}{arg2}" # pyright: ignore[reportOptionalOperand]
    save_json(arg3, chunks_output)
    save_jsonl(arg3, chunks_jsonl_output)

def main():
    """Main execution function."""
    try:
        # Setup
        base_dir = find_project_root()
        setup_directories(base_dir)
        log_file = base_dir / "logs" / "banka_dokumanlari_preparation_excel.log"
        global logger
        logger = setup_logging(log_file)
        
        logger.info("Starting Banka Dokümanları Preparation - Excel Only")
        logger.info(f"Base directory: {base_dir}")
        logger.info(f"Raw directory: {BANKA_DOKUMANLARI_RAW_DIR}")
        
        # Process all Excel documents in raw directory
        if not BANKA_DOKUMANLARI_RAW_DIR.exists(): # pyright: ignore[reportOptionalMemberAccess]
            logger.error(f"Raw directory does not exist: {BANKA_DOKUMANLARI_RAW_DIR}")
            return
        
        # Find all Excel documents
        excel_extensions = ['.xlsx', '.xls']
        excel_documents = [f for f in BANKA_DOKUMANLARI_RAW_DIR.iterdir()  # pyright: ignore[reportOptionalMemberAccess]
                          if f.is_file() and f.suffix.lower() in excel_extensions]
        
        if not excel_documents:
            logger.warning("No Excel documents found in raw directory")
            return
        
        logger.info(f"Found {len(excel_documents)} Excel documents to process")
        
        for doc_path in excel_documents:
            try:
                process_excel_document(doc_path)
            except Exception as e:
                logger.error(f"Failed to process {doc_path.name}: {e}")
                continue
        
        logger.info("All Excel documents processed successfully")
        
    except Exception as e:
        logger.error(f"Error in main execution: {e}")
        raise

if __name__ == "__main__":
    main()
